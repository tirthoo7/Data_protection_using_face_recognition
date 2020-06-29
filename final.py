from os import path
import PIL.Image
# pip install opencv-contrib-python
import tkinter as tk
from tkinter import * 
from tkinter import ttk,PhotoImage
# import Pillow as PIL
from PIL import Image
# import Image
from tkinter import messagebox
import os
#import mysql.connector
import cv2
import numpy as np

import xlrd 
import openpyxl 
import xlsxwriter 
import xlwt 
from xlwt import Workbook 
detector=cv2.CascadeClassifier('XML/haarcascade_frontalface_default.xml')
recognizer = cv2.face.LBPHFaceRecognizer_create()

def login():
	
	login=Tk()
	login.title("Admin Login")

	main_frame_container=Frame(login,bg='#E8DAEF')
	main_frame_container.pack(fill=BOTH,expand=True )


	username_label=tk.Label(main_frame_container,text="Username",bg='#F1948A')
	username=tk.StringVar()
	username_textField=tk.Entry(main_frame_container,width=20,textvariable=username,
								bg='#D6EAF8')

	password=StringVar()
	password_label=tk.Label(main_frame_container,text="Password",bg='#F1948A')
	password_textField=tk.Entry(main_frame_container,width=20,textvariable=password,show="*",
								bg='#D6EAF8')

	username_label.pack()
	username_textField.pack()

	password_label.pack()
	password_textField.pack()

	Submit_bt=tk.Button(main_frame_container,text="Login")
	Submit_bt.pack()

	var=StringVar()
	# mess=tk.Message(main_frame_container,textvariable=var,bg='#73C6B6',padx=100,width=100)  
	# mess.pack()

	def verify(event):

		un=username.get()
		pw=password.get()
		#print(f"Username:-{un}")
		#print(f"Password:-{pw}")
	   # if(un=='tirth' and pw=='007') :
		if(True):
			var.set("Welcome")
			login.destroy()
			Face_Recog()
			
		else:
			var.set("Invalid")

	Submit_bt.bind('<Button-1>',verify)        

	login.geometry('500x500')
	#login.iconbitmap(default='img/png_icon.png')
	# login.mainloop()


def main_login():
	main=Tk()
	main.title("Welcome")
	main.geometry("300x200")
	main_container=Frame(main,bg='#E8DAEF')
	main_container.pack(fill=BOTH,expand=True )
	Recog_b = Button(main_container, text='Recognition',  
					  command=recog, width=15, bg="#FFF59D").pack(fill=BOTH)
	Admin_b= Button(main_container, text='Admin',  
					 command=login, width=15, bg="#FFF59D").pack(fill=BOTH)
	main.mainloop()

def get():
	db=[]
	if(not(path.exists("./maping.xlsx"))):
		return db
		# Reading an excel file using Python 

	# Give the location of the file 
	loc = ('./maping.xlsx') 

	# To open Workbook 
	wb = xlrd.open_workbook(loc) 
	sheet = wb.sheet_by_index(0) 

	# For row 0 and column 0 
	number_of_rows=sheet.nrows
	#print(number_of_rows)
	#total_faces=number_of_rows-1
	#print(total_faces)
	
	for row in range(0,number_of_rows):
		
	   # print(sheet.row_values(row))#list of values
	   # print(sheet.row_values(row)[0])#face id
	   # print(sheet.row_values(row)[1])#face name
		face_id=sheet.row_values(row)[0]
		face_name=sheet.row_values(row)[1]
		data={'id':face_id,
		 'name':face_name}
		db.append(data)
	return db


# In[46]:


def createDb():
	workbook = xlsxwriter.Workbook('./maping.xlsx')
	workbook.close() 
	# It is neccessary to close file to save it
	
	
# createDb()



def store_db(Face_id,Face_name):
	if(not(path.exists("./maping.xlsx"))):
		createDb()
	


	loc = ('./maping.xlsx') 

	# To open Workbook 
	wb = xlrd.open_workbook(loc) 
	sheet = wb.sheet_by_index(0) 
	current_row=sheet.nrows+1
	
 
	wbkName = './maping.xlsx'
	wbk = openpyxl.load_workbook(wbkName)
	sheet = wbk.active 
	c=sheet.cell(row=current_row, column=1)
	c.value=str(Face_id)
	c=sheet.cell(row=current_row, column=2)
	c.value=str(Face_name)

	wbk.save('./maping.xlsx')
	wbk.close
 
def getImagesAndLabels(path):
	# detector= cv2.CascadeClassifier("XML/haarcascade_frontalface_default.xml");
	#get the path of all the files in the folder
	imagePaths=[os.path.join(path,f) for f in os.listdir(path)] 
	#create empth face list
	faceSamples=[]
	#create empty ID list
	Ids=[]
	#now looping through all the image paths and loading the Ids and the images
	for imagePath in imagePaths:
		#must be .jpg
		if(os.path.split(imagePath)[-1].split(".")[-1]!='jpg'):
			continue
		#loading the image and converting it to gray scale
		pilImage=PIL.Image.open(imagePath).convert('L')
		#Now we are converting the PIL image into numpy array
		imageNp=np.array(pilImage,'uint8')
		#getting the Id from the image
		Id=int(os.path.split(imagePath)[-1].split(".")[1])
		# extract the face from the training image sample
		faces=detector.detectMultiScale(imageNp)
		#If a face is there then append that in the list as well as Id of it
		for (x,y,w,h) in faces:
			faceSamples.append(imageNp[y:y+h,x:x+w])
			Ids.append(Id)
	return faceSamples,Ids


def capImg():
	
	try:
		
		take_ip_window=Toplevel()
		def f(event):

			if (face_id_ip.get().isdecimal()==False):
				print("ID must be Int")
				take_ip_window.destroy()
				return
			face_id=face_id_ip.get()
			face_name=face_name_ip.get()
			
			take_ip_window.destroy()
			# detector=cv2.CascadeClassifier('XML/haarcascade_frontalface_default.xml')
			sampleNum=0
			cam = cv2.VideoCapture(0)
			while(True):
				ret, img = cam.read()

				gray = cv2.cvtColor(img, cv2.COLOR_BGR2GRAY)
				faces = detector.detectMultiScale(gray, 1.3, 5)
				for (x,y,w,h) in faces:
					cv2.rectangle(img,(x,y),(x+w,y+h),(255,0,0),2)

					#incrementing sample number 
					sampleNum=sampleNum+1
					#saving the captured face in the dataset folder
					cv2.imwrite("dataset_of_faces/User."+face_id+'.'+ str(sampleNum) + ".jpg", gray[y:y+h,x:x+w])

					cv2.imshow('frame',img)
				#wait for 100 miliseconds 
				k = cv2.waitKey(10) & 0xff # Press 'ESC' for exiting video
				if k == 27:
					break
				# break if the sample number is morethan 20
				elif sampleNum>30:
					store_db(face_id,face_name)
					break
			cam.release()
			cv2.destroyAllWindows()
	except:
		print("Capturing is Not Working")
		cam.release()
		cv2.destroyAllWindows()
	
	
	take_ip_frame=Frame(take_ip_window,bg='#66FF99')
	take_ip_frame.pack(fill=BOTH,expand=True )
	
	face_id_label=Label(take_ip_frame,text=" enter user id(int) :-")
	face_id_label.pack()
	
	face_id_ip=StringVar()
	face_id_Entry=Entry(take_ip_frame,textvariable=face_id_ip,bg='#CCFFFF')
	face_id_Entry.pack()
	
	face_name_label=Label(take_ip_frame,text=" enter user name :-")
	face_name_label.pack()
	
	face_name_ip=StringVar()
	face_name_Entry=Entry(take_ip_frame,textvariable=face_name_ip,bg='#CCFFFF')
	face_name_Entry.pack()
	
	Submit_bt=tk.Button(take_ip_frame,text="Submit",bg='#FF9999')
	Submit_bt.pack()
	Submit_bt.bind('<Button-1>',f)  
	
	
	take_ip_window.geometry('500x500')
	take_ip_window.mainloop()

#train Image
def trn():
	# recognizer = cv2.face.LBPHFaceRecognizer_create()
	# detector= cv2.CascadeClassifier("XML/haarcascade_frontalface_default.xml");
	faces,Ids = getImagesAndLabels('dataset_of_faces/')
	recognizer.train(faces, np.array(Ids))
	recognizer.save('trainner.yml')
	print("Trainig Completed")


# In[44]:


#image is trained and we will now recognize
def recog():
	match_flag=False
	# recognizer = cv2.face.LBPHFaceRecognizer_create()
	if(not(path.exists("trainner.yml"))):
		print("yml file is not found ")
		print("First Train the face")
		return -1


	recognizer.read('trainner.yml')
	
	# faceCascade = cv2.CascadeClassifier("XML/haarcascade_frontalface_default.xml");
	font = cv2.FONT_HERSHEY_SIMPLEX
	cam = cv2.VideoCapture(0)
	while True:
		ret, im =cam.read()
		
		user_name="Unknown"    
		gray=cv2.cvtColor(im,cv2.COLOR_BGR2GRAY)
		faces=detector.detectMultiScale(gray, 1.2,5)
		
		for(x,y,w,h) in faces:
			cv2.rectangle(im,(x,y),(x+w,y+h),(225,0,0),2)
			Id, conf = recognizer.predict(gray[y:y+h,x:x+w])
#             print("Detected:"+"ID -"+str(Id))
			# confidence-> loss (lesser is better)
			match_flag=False
			user_name="Unknown" 
			db=get()
			print(db)
			for data in db:
				
				if(str(Id)==str(data['id']) and conf<70 and conf>40 ):
					match_flag=True
					user_name=data['name']
					print("Face Found")
					cv2.putText(im, str(user_name), (x+5,y-5), font, 1, (255,255,255), 2)
					cv2.putText(im, str(conf), (x+5,y+h-5), font, 1, (255,255,0), 1) 
				else:
					user_name="Unknown"
					match_flag=False

			cv2.putText(im, str(user_name), (x+5,y-5), font, 1, (255,255,255), 2)
			cv2.putText(im, str(conf), (x+5,y+h-5), font, 1, (255,255,0), 1) 
			
				
		   
			
		cv2.imshow('im',im) 
		k = cv2.waitKey(10) & 0xff # Press 'ESC' for exiting video
		if k == 27:
			break
		elif (k==13):
			break

			
				

	cam.release()
	cv2.destroyAllWindows()
	if(match_flag):
		show_notes()
   
def Face_Recog():

	face_recog=Tk()
	face_recog.title("Face Enrollement")
	face_recog.geometry('420x520')
	
	main_frame_container=Frame(face_recog,bg='#B2EBF2')
	main_frame_container.pack(fill=BOTH,expand=True )
	
	capb_img = Button(main_frame_container, text='Capture Image',  
					  command=capImg, width=15, bg="#FFF59D").pack()
	trn_img = Button(main_frame_container, text='Train Image',  
					 command=trn, width=15, bg="#FFF59D").pack()
	rc_img = Button(main_frame_container, text='Recognize Image',  
					command=recog, width=15, bg="#FFF59D").pack()

	clear_data = Button(main_frame_container, text='Clear Data',  
					command=clearData, width=15, bg="#FFF59D").pack()

# class Notes:

# 	data=""
# #     def __init__(self,data):
# #         self.data=data
# 	def appendData(self,s):
#   		self.data=self.data+"\n"+s;

# 	def getData(self):
# 		return self.data

# myNotes=Notes()

def clearData():
	print("Clearing")

	if os.path.exists("./maping.xlsx"):
		os.remove("./maping.xlsx")
	else:
		pass

	mydir='./dataset_of_faces'	
	filelist =os.listdir(mydir) 
	for f in filelist:
		os.remove(os.path.join(mydir, f))

	if os.path.exists("./trainner.yml"):
		os.remove("./trainner.yml")
	else:
		pass	

	print("Cleared")	


def show_notes():
	def saveDetails():
		# print(text.get("1.0",END))
		reader = open('./data.txt','w')
		try:
			reader.write(text.get("1.0",END))
		finally:
			reader.close()

	
	notes=Tk()
	notes.title("Your Notes")
	text=tk.Text(notes,height=15,width=30,bg="#EAF2F8")
	reader = open('./data.txt','r')
	try:
		text.insert(tk.INSERT,reader.read())
	finally:
		reader.close()


	save_butt = tk.Button(notes, text ="Save", command = saveDetails,bg="#A3E4D7")
	

	# text.insert(tk.INSERT,"YOUR ATM Data1")
	# text.insert(tk.INSERT,"YOUR ATM Data2s")

	text.pack()
	save_butt.pack()	

	notes.mainloop() 




# show_notes()
# login()

try:
	main_login()
  # login()
except:
	print("Something went wrong")
finally:
	cv2.destroyAllWindows()
